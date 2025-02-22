import sys
from openpyxl import load_workbook

def clear_excel_data(file_path, keep_headers=False):
    try:
        # Load the workbook and select the active sheet
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Clear the sheet data
        if keep_headers:
            # Clear all rows except the header
            for row in sheet.iter_rows(min_row=2):  # Start from the second row
                for cell in row:
                    cell.value = None
        else:
            # Clear all rows
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None

        # Save the workbook
        workbook.save(file_path)
        print(f"Data cleared successfully in '{file_path}'!")
    except Exception as e:
        print(f"Error clearing data: {e}")

if __name__ == "__main__":
    # Get arguments from the command line
    if len(sys.argv) < 2:
        print("Usage: python clear_excel_data.py <file_path> [--keep-headers]")
        sys.exit(1)

    file_path = sys.argv[1]
    keep_headers = "--keep-headers" in sys.argv

    clear_excel_data(file_path, keep_headers)

